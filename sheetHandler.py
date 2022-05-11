from os.path import exists
import openpyxl
from thompcoutils.log_utils import get_logger
from siteData import SiteData


class SheetHandler:
    ALPHABET = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    HEADING_ROW = 1
    col = 1
    NUM_COL = col
    col += 1
    ACCOUNT_COL = col
    col += 1
    OWNER_COL = col
    col += 1
    MAILING_ADDRESS_COL = col
    col += 1
    SITE_ADDRESS_COL = col
    col += 1
    NOTES_COL = col
    col += 1

    def __init__(self, file_name):
        """
        Initialize the class with the file name.
        param file_name:
        """
        self.file_name = file_name
        create = not exists(file_name)
        if create:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.cell(row=SheetHandler.HEADING_ROW, column=SheetHandler.NUM_COL).value = '#'
            ws.cell(row=SheetHandler.HEADING_ROW, column=SheetHandler.ACCOUNT_COL).value = 'Account'
            ws.cell(row=SheetHandler.HEADING_ROW, column=SheetHandler.OWNER_COL).value = 'Owner'
            ws.cell(row=SheetHandler.HEADING_ROW, column=SheetHandler.MAILING_ADDRESS_COL).value = \
                'Mailing Address'
            ws.cell(row=SheetHandler.HEADING_ROW, column=SheetHandler.SITE_ADDRESS_COL).value = 'Site Address'
            ws.cell(row=SheetHandler.HEADING_ROW, column=SheetHandler.NOTES_COL).value = 'Notes'
            ws.cell(row=SheetHandler.HEADING_ROW + 1, column=SheetHandler.NUM_COL).value = '=1'
            for x in range(2, 20):
                ws.cell(row=x + 1, column=SheetHandler.NUM_COL).value = \
                    '={}{}+1'.format(SheetHandler.ALPHABET[SheetHandler.NUM_COL - 1], x)
            wb.save(file_name)
            exit(0)
        else:
            self.values = None
            self.wb = openpyxl.load_workbook(file_name)

    def get_accounts(self):
        logger = get_logger()
        if self.values is None:
            sheet = self.wb.active
            row = SheetHandler.HEADING_ROW + 1
            self.values = []
            while True:
                account = sheet.cell(row=row, column=SheetHandler.ACCOUNT_COL).value
                if account is None:
                    logger.debug('End of sheet')
                    break
                else:
                    name = sheet.cell(row=row, column=SheetHandler.OWNER_COL).value
                    mailing_address = sheet.cell(row=row, column=SheetHandler.MAILING_ADDRESS_COL).value
                    site_address = sheet.cell(row=row, column=SheetHandler.SITE_ADDRESS_COL).value
                    notes = sheet.cell(row=row, column=SheetHandler.NOTES_COL).value
                    logger.debug('Spreadsheet account:{}, site_address:{}, notes:{}'.
                                 format(account, site_address, notes))
                    self.values.append(SiteData(account,
                                                name,
                                                mailing_address,
                                                site_address,
                                                notes))
                    row += 1
        return self.values

    def update_account(self, site_data):
        sheet = self.wb.active
        row = SheetHandler.HEADING_ROW + 1
        while True:
            account_cell = sheet.cell(row=row, column=SheetHandler.ACCOUNT_COL)
            if account_cell.value is None:
                raise Exception("Account not found")
            else:
                if account_cell.value == site_data.account:
                    sheet.cell(row=row, column=SheetHandler.OWNER_COL).value = site_data.owner
                    sheet.cell(row=row, column=SheetHandler.MAILING_ADDRESS_COL).value = \
                        site_data.mailing_address
                    sheet.cell(row=row, column=SheetHandler.SITE_ADDRESS_COL).value = \
                        site_data.site_address
                    self.wb.save(self.file_name)
                    return
            row += 1
